from django.db import models

class TeamRegistration(models.Model):
    primary_name = models.CharField(max_length=100)
    teammate_name = models.CharField(max_length=100, blank=True, null=True)
    
    primary_email = models.EmailField(unique=True)
    teammate_email = models.EmailField(blank=True, null=True)
    
    primary_mobile = models.CharField(max_length=15)
    teammate_mobile = models.CharField(max_length=15, blank=True, null=True)
    
    team_name = models.CharField(max_length=100)
    institution = models.CharField(max_length=200)
    
    experience = models.TextField(blank=True, null=True)
    
    # Cloudinary URL for the payment screenshot (uploaded at registration time)
    payment_screenshot = models.URLField(max_length=500, blank=True, null=True)
    
    # Admin approval flag
    is_approved = models.BooleanField(default=False)
    
    # Generated credential mapping
    generated_username = models.CharField(max_length=50, blank=True, null=True)
    
    # Merch Fields
    add_merch = models.BooleanField(default=False)
    primary_tshirt_size = models.CharField(max_length=5, blank=True, null=True)
    teammate_tshirt_size = models.CharField(max_length=5, blank=True, null=True)
    
    # Delegate Assignment Fields
    debate_topic = models.CharField(max_length=500, default='Waiting for Assignment...', blank=True, null=True)
    stance = models.CharField(max_length=100, default='Pending', blank=True, null=True)
    debate_date = models.CharField(max_length=100, default='TBD', blank=True, null=True)
    debate_time = models.CharField(max_length=100, default='TBD', blank=True, null=True)
    classroom = models.CharField(max_length=100, default='TBD', blank=True, null=True)
    
    # Optional Referral Code
    referral_code = models.CharField(max_length=100, blank=True, null=True)

    def __str__(self):
        return f"{self.team_name} - {self.primary_name}"

    def save(self, *args, **kwargs):
        try:
            if self.pk:
                orig = TeamRegistration.objects.get(pk=self.pk)
                was_approved = orig.is_approved
            else:
                was_approved = False
        except TeamRegistration.DoesNotExist:
            was_approved = False

        newly_approved = (not was_approved and self.is_approved)
        should_sync_excel = self.is_approved

        plain_password = None
        if newly_approved and not self.generated_username:
            import string, random
            from django.contrib.auth.models import User
            # Crendential Gen
            base_login_id = self.team_name.replace(' ', '_').lower()
            login_id = base_login_id
            counter = 1
            while User.objects.filter(username=login_id).exists():
                login_id = f"{base_login_id}_{counter}"
                counter += 1
            
            plain_password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
            User.objects.create_user(username=login_id, email=self.primary_email, password=plain_password)
            self.generated_username = login_id

        super().save(*args, **kwargs)

        # Automatically mail credentials if newly approved by spawning a detached subprocess.
        # This keeps HTTP requests lightning-fast and prevents Gunicorn from killing worker processes.
        if newly_approved and plain_password:
            import subprocess
            import sys
            import os
            from django.conf import settings
            
            manage_py_path = os.path.join(settings.BASE_DIR, 'manage.py')
            teammate_email_arg = self.teammate_email or 'None'
            
            try:
                subprocess.Popen([
                    sys.executable,
                    manage_py_path,
                    'send_approval_email',
                    self.primary_email,
                    teammate_email_arg,
                    self.team_name,
                    self.generated_username,
                    plain_password
                ], close_fds=True)
                print(f"[Subprocess] Spawned background email sender process for {self.primary_email}.")
            except Exception as e:
                print(f"[Subprocess] Failed to spawn email process: {e}")

        if should_sync_excel:
            try:
                import os, openpyxl
                from openpyxl import Workbook
                from django.conf import settings
            except ImportError:
                print("WARNING: openpyxl is not installed. Excel sync skipped. Run: pip install openpyxl")
                return

            excel_path = os.path.join(settings.BASE_DIR, 'infacto_participants.xlsx')

            try:
                if os.path.exists(excel_path):
                    wb = openpyxl.load_workbook(excel_path)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.append([
                        'Team Name', 'Primary Name', 'Primary Email', 'Primary Mobile',
                        'Teammate Name', 'Teammate Email', 'Teammate Mobile',
                        'Institution', 'Experience', 'Login ID', 'Password',
                        'Merch Opt-In', 'Primary Size', 'Teammate Size',
                        'Debate Topic', 'Stance', 'Date', 'Time', 'Classroom',
                        'Referral Code'
                    ])

                team_row = None
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row=row, column=1).value == self.team_name:
                        team_row = row
                        break

                merch_status = "Yes" if self.add_merch else "No"

                if newly_approved and not team_row:
                    ws.append([
                        self.team_name, self.primary_name, self.primary_email, self.primary_mobile,
                        self.teammate_name, self.teammate_email, self.teammate_mobile,
                        self.institution, self.experience, self.generated_username, plain_password or "Hidden",
                        merch_status, self.primary_tshirt_size or "-", self.teammate_tshirt_size or "-",
                        self.debate_topic, self.stance, self.debate_date, self.debate_time, self.classroom,
                        self.referral_code or "-"
                    ])
                elif team_row:
                    ws.cell(row=team_row, column=12, value=merch_status)
                    ws.cell(row=team_row, column=13, value=self.primary_tshirt_size or "-")
                    ws.cell(row=team_row, column=14, value=self.teammate_tshirt_size or "-")
                    ws.cell(row=team_row, column=15, value=self.debate_topic)
                    ws.cell(row=team_row, column=16, value=self.stance)
                    ws.cell(row=team_row, column=17, value=self.debate_date)
                    ws.cell(row=team_row, column=18, value=self.debate_time)
                    ws.cell(row=team_row, column=19, value=self.classroom)

                wb.save(excel_path)
            except Exception as e:
                print(f"Skipping Excel Background Sync: File is likely locked by another program. Error: {str(e)}")

